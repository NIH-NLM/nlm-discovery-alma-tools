import argparse
import os
import re
import time
from pathlib import Path
from typing import Any, cast
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import openpyxl
import requests
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert source URLs to watch-format URLs, follow redirects, and write "
            "results into an output XLSX file."
        )
    )
    parser.add_argument("input_xlsx", help="Path to input XLSX file")
    parser.add_argument(
        "-o",
        "--output-xlsx",
        help="Path to output XLSX file (default: <input_stem>_output.xlsx)",
    )
    parser.add_argument(
        "--source-column",
        default="Uniform Resource Identifier",
        help="Header name for the source URL column",
    )
    parser.add_argument(
        "--converted-column",
        default="Converted URL",
        help="Header name for converted URL results",
    )
    parser.add_argument(
        "--redirected-column",
        default="Redirected URL",
        help="Header name for redirected URL results",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=15,
        help="HTTP timeout in seconds for each URL request",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.5,
        help="Delay in seconds between URL requests",
    )
    parser.add_argument(
        "--save-every",
        type=int,
        default=50,
        help="Save output workbook every N rows",
    )
    parser.add_argument(
        "--alma-api-key",
        help="Optional Alma API key (used only for Alma API URLs)",
    )
    parser.add_argument(
        "--alma-api-key-file",
        help=(
            "Optional path to text file containing alma_api_key = \"...\"; "
            "used only for Alma API URLs"
        ),
    )
    return parser.parse_args()


def find_column_index(sheet: Worksheet, header_name: str):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == header_name:
            return col
    return None


def convert_url_value(raw_value: Any) -> str:
    if raw_value is None or str(raw_value).strip() == "":
        return "manual cleanup"

    url_str = str(raw_value).strip()

    if "watch=" in url_str:
        return url_str

    if "launch.asp?" in url_str:
        return re.sub(r"launch\.asp\?", "watch=", url_str)

    return "manual cleanup"


def load_alma_api_key_from_file(file_path: str) -> str | None:
    try:
        with open(file_path, "r", encoding="utf-8") as handle:
            content = handle.read()
    except OSError:
        return None

    match = re.search(r"alma_api_key\s*=\s*['\"]([^'\"]+)['\"]", content)
    if match:
        return match.group(1).strip()

    # Fallback for files that contain only the key on one line.
    for line in content.splitlines():
        candidate = line.strip()
        if candidate and not candidate.startswith("#"):
            return candidate.strip("'\"")
    return None


def is_alma_api_url(url: str) -> bool:
    parsed = urlparse(url)
    return (
        parsed.scheme in {"http", "https"}
        and parsed.netloc.endswith("hosted.exlibrisgroup.com")
        and "/almaws/" in parsed.path
    )


def with_alma_api_key(url: str, alma_api_key: str | None) -> str:
    if not alma_api_key or not is_alma_api_url(url):
        return url

    parsed = urlparse(url)
    query_items = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query_items["apikey"] = alma_api_key
    new_query = urlencode(query_items)
    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            parsed.params,
            new_query,
            parsed.fragment,
        )
    )


def get_redirected_url(
    url: str,
    timeout: float,
    session: requests.Session,
    alma_api_key: str | None,
) -> str:
    try:
        request_url = with_alma_api_key(url, alma_api_key)
        response = session.get(request_url, timeout=timeout, allow_redirects=True)
        return response.url if response.url != url else "no redirect"
    except requests.RequestException:
        return "no redirect"


def derive_output_path(input_path: Path, output_path_arg: str | None) -> Path:
    if output_path_arg:
        return Path(output_path_arg)
    return input_path.with_name(f"{input_path.stem}_output.xlsx")


def set_cell_value(sheet: Worksheet, row: int, col: int, value: str) -> None:
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_xlsx)
    output_path = derive_output_path(input_path, args.output_xlsx)
    alma_api_key = args.alma_api_key or os.getenv("ALMA_API_KEY")
    if args.alma_api_key_file and not alma_api_key:
        alma_api_key = load_alma_api_key_from_file(args.alma_api_key_file)

    if not input_path.exists():
        print(f"Error: input file not found: {input_path}")
        return 1

    wb = openpyxl.load_workbook(input_path)
    sheet = cast(Worksheet, wb.active)

    source_col = find_column_index(sheet, args.source_column)
    if source_col is None:
        print(f"Error: source column '{args.source_column}' not found in row 1.")
        return 1

    converted_col = sheet.max_column + 1
    redirected_col = converted_col + 1
    set_cell_value(sheet, 1, converted_col, args.converted_column)
    set_cell_value(sheet, 1, redirected_col, args.redirected_column)

    total_rows = max(sheet.max_row - 1, 0)
    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print(f"Rows:   {total_rows}")
    if alma_api_key:
        print("Alma API key mode: enabled (Alma API URLs only)")
    print("-" * 60)

    print("STEP 1: Building converted URLs")
    for row in range(2, sheet.max_row + 1):
        original = sheet.cell(row=row, column=source_col).value
        converted = convert_url_value(original)
        set_cell_value(sheet, row, converted_col, converted)

    print("STEP 2: Following redirects")
    session = requests.Session()
    session.headers.update(
        {"User-Agent": "url-redirect-template/1.0 (+https://github.com/)"}
    )

    for row in range(2, sheet.max_row + 1):
        progress = row - 1
        converted_value = sheet.cell(row=row, column=converted_col).value

        if converted_value is None or converted_value == "manual cleanup":
            set_cell_value(sheet, row, redirected_col, "no redirect")
            continue

        urls = [u.strip() for u in str(converted_value).split(",") if u.strip()]
        redirected_urls = []

        for url in urls:
            if "watch=" in url:
                redirected = get_redirected_url(url, args.timeout, session, alma_api_key)
                redirected_urls.append(redirected)
                time.sleep(args.delay)
            else:
                redirected_urls.append("no redirect")

        if progress % args.save_every == 0:
            wb.save(output_path)
            print(f"Saved progress at row {row} of {sheet.max_row}")

        set_cell_value(sheet, row, redirected_col, ", ".join(redirected_urls))

    wb.save(output_path)
    print("-" * 60)
    print("Done.")
    print(f"Saved output workbook: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
